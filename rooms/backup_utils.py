import io
from django.utils import timezone
from django.http import HttpResponse
from .models import Booking, Room, RoomType, DataBackup, CustomUser

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def export_bookings_to_excel(start_date=None, end_date=None):
    """
    Export bookings data to Excel format
    Returns: BytesIO object containing Excel file data
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("Excel libraries (openpyxl, pandas) not available")
    # Get bookings data
    bookings_query = Booking.objects.select_related('room', 'room__room_type', 'created_by')
    
    if start_date and end_date:
        bookings_query = bookings_query.filter(
            check_in_date__gte=start_date,
            check_out_date__lte=end_date
        )
    
    bookings = bookings_query.order_by('-created_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    
    # Define headers
    headers = [
        'ID', 'Room Number', 'Room Type', 'Guest Name', 'Guest Contact',
        'Check In Date', 'Check Out Date', 'Total Amount', 'Paid Amount',
        'Status', 'Payment Status', 'Notes', 'Created By', 'Created At', 'Updated At'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_num, booking in enumerate(bookings, 2):
        ws.cell(row=row_num, column=1, value=booking.id)
        ws.cell(row=row_num, column=2, value=booking.room.room_number)
        ws.cell(row=row_num, column=3, value=booking.room.room_type.get_name_display())
        ws.cell(row=row_num, column=4, value=booking.guest_name)
        ws.cell(row=row_num, column=5, value=booking.guest_contact)
        ws.cell(row=row_num, column=6, value=booking.check_in_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=7, value=booking.check_out_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=8, value=float(booking.total_amount))
        ws.cell(row=row_num, column=9, value=float(booking.paid_amount))
        ws.cell(row=row_num, column=10, value=booking.get_status_display())
        ws.cell(row=row_num, column=11, value=booking.get_payment_status_display())
        ws.cell(row=row_num, column=12, value=booking.notes)
        ws.cell(row=row_num, column=13, value=booking.created_by.username if booking.created_by else '')
        ws.cell(row=row_num, column=14, value=booking.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=15, value=booking.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Add rooms data sheet
    rooms_ws = wb.create_sheet(title="Rooms")
    room_headers = ['ID', 'Room Number', 'Room Type', 'Room Type Display', 'Is Active', 
                   'Weekday Rate', 'Weekend Rate']
    
    # Write room headers
    for col_num, header in enumerate(room_headers, 1):
        cell = rooms_ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Write room data
    rooms = Room.objects.select_related('room_type').order_by('room_number')
    for row_num, room in enumerate(rooms, 2):
        rooms_ws.cell(row=row_num, column=1, value=room.id)
        rooms_ws.cell(row=row_num, column=2, value=room.room_number)
        rooms_ws.cell(row=row_num, column=3, value=room.room_type.name)
        rooms_ws.cell(row=row_num, column=4, value=room.room_type.get_name_display())
        rooms_ws.cell(row=row_num, column=5, value=room.is_active)
        rooms_ws.cell(row=row_num, column=6, value=float(room.room_type.base_weekday_rate))
        rooms_ws.cell(row=row_num, column=7, value=float(room.room_type.base_weekend_rate))
    
    # Auto-adjust column widths
    for ws_sheet in [ws, rooms_ws]:
        for column in ws_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def import_bookings_from_excel(excel_file, user=None):
    """
    Import bookings from Excel file
    Returns: dict with success status, message, and imported count
    """
    if not EXCEL_AVAILABLE:
        return {
            'success': False,
            'message': 'Excel libraries not available for import functionality',
            'imported_count': 0
        }
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name='Bookings')
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Get or create room
                room_number = str(row['Room Number']).strip()
                room = Room.objects.filter(room_number=room_number).first()
                
                if not room:
                    errors.append(f"Row {index + 2}: Room {room_number} not found")
                    continue
                
                # Check if booking already exists (by guest name, room, and dates)
                existing_booking = Booking.objects.filter(
                    room=room,
                    guest_name=str(row['Guest Name']).strip(),
                    check_in_date=pd.to_datetime(row['Check In Date']).date(),
                    check_out_date=pd.to_datetime(row['Check Out Date']).date()
                ).first()
                
                if existing_booking:
                    # Update existing booking
                    existing_booking.guest_contact = str(row.get('Guest Contact', '')).strip()
                    existing_booking.total_amount = row['Total Amount']
                    existing_booking.paid_amount = row.get('Paid Amount', 0)
                    existing_booking.status = get_status_from_display(str(row.get('Status', 'PENCIL')))
                    existing_booking.payment_status = get_payment_status_from_display(str(row.get('Payment Status', 'UNPAID')))
                    existing_booking.notes = str(row.get('Notes', '')).strip()
                    existing_booking.save()
                    imported_count += 1
                else:
                    # Create new booking
                    booking = Booking.objects.create(
                        room=room,
                        guest_name=str(row['Guest Name']).strip(),
                        guest_contact=str(row.get('Guest Contact', '')).strip(),
                        check_in_date=pd.to_datetime(row['Check In Date']).date(),
                        check_out_date=pd.to_datetime(row['Check Out Date']).date(),
                        total_amount=row['Total Amount'],
                        paid_amount=row.get('Paid Amount', 0),
                        status=get_status_from_display(str(row.get('Status', 'PENCIL'))),
                        payment_status=get_payment_status_from_display(str(row.get('Payment Status', 'UNPAID'))),
                        notes=str(row.get('Notes', '')).strip(),
                        created_by=user
                    )
                    imported_count += 1
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        if errors:
            return {
                'success': False,
                'message': f"Imported {imported_count} bookings with {len(errors)} errors",
                'errors': errors,
                'imported_count': imported_count
            }
        else:
            return {
                'success': True,
                'message': f"Successfully imported {imported_count} bookings",
                'imported_count': imported_count
            }
            
    except Exception as e:
        return {
            'success': False,
            'message': f"Error reading Excel file: {str(e)}",
            'imported_count': 0
        }


def get_status_from_display(display_value):
    """Convert display value back to database value"""
    status_map = {
        'Pencil Booked': 'PENCIL',
        'Confirmed': 'CONFIRMED',
        'Checked In': 'CHECKED_IN',
        'No Show': 'NO_SHOW',
        'Cancelled': 'CANCELLED',
    }
    return status_map.get(display_value, 'PENCIL')


def get_payment_status_from_display(display_value):
    """Convert payment status display value back to database value"""
    status_map = {
        'Unpaid': 'UNPAID',
        'Partial Payment': 'PARTIAL',
        'Fully Paid': 'PAID',
    }
    return status_map.get(display_value, 'UNPAID')


def create_backup_record(backup_type='AUTO', user=None, file_data=None, notes=''):
    """
    Create a backup record in the database
    """
    now = timezone.now()
    file_name = f"hotel_backup_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    if file_data is None:
        file_data = export_bookings_to_excel().getvalue()
    
    booking_count = Booking.objects.count()
    room_count = Room.objects.count()
    
    backup = DataBackup.objects.create(
        backup_type=backup_type,
        created_by=user,
        file_name=file_name,
        file_data=file_data,
        booking_count=booking_count,
        room_count=room_count,
        notes=notes
    )
    
    return backup


def cleanup_old_backups():
    """
    Clean up old automatic backups according to 24-hour rotation rule
    Keep only the last 144 backups (10 min intervals for 24 hours)
    """
    cutoff_time = timezone.now() - timezone.timedelta(hours=24)
    
    # Get automatic backups older than 24 hours
    old_backups = DataBackup.objects.filter(
        backup_type='AUTO',
        created_at__lt=cutoff_time
    ).order_by('-created_at')
    
    # Keep manual backups and imports, only delete auto backups
    deleted_count = 0
    for backup in old_backups:
        backup.delete()
        deleted_count += 1
    
    return deleted_count


def generate_backup_filename():
    """Generate a standardized backup filename"""
    now = timezone.now()
    return f"hotel_backup_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"